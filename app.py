# =========================
# 環境変数ロード
# =========================
from dotenv import load_dotenv
load_dotenv()

# =========================
# 標準ライブラリ
# =========================
import os
import io
import smtplib
from datetime import datetime, timedelta

# =========================
# サードパーティ
# =========================
from flask import Flask, render_template, request, redirect, session, jsonify, send_file
from supabase import create_client
from openpyxl import Workbook
from email.mime.text import MIMEText
from email.header import Header
from email.utils import formataddr

# =========================
# Flask設定
# =========================
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY")
app.permanent_session_lifetime = timedelta(days=7)

# =========================
# Supabase接続
# =========================
supabase = create_client(
    os.getenv("SUPABASE_URL"),
    os.getenv("SUPABASE_KEY")
)

def format_time_range(time_str):
    t = time_str[:5]  # "09:30" に統一

    start = datetime.strptime(t, "%H:%M")
    end = start + timedelta(minutes=30)

    return f"{start.strftime('%H:%M')}～{end.strftime('%H:%M')}"

def send_admin_mail(subject, body):

    ADMIN_EMAIL = os.getenv("ADMIN_EMAIL")
    SMTP_USER = os.getenv("SMTP_USER")
    SMTP_PASS = os.getenv("SMTP_PASS")

    if not SMTP_USER or not SMTP_PASS or not ADMIN_EMAIL:
        print("SMTP環境変数が不足しています")
        return

    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"] = SMTP_USER
    msg["To"] = ADMIN_EMAIL

    try:
        with smtplib.SMTP("smtp.gmail.com", 587, timeout=10) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()

            server.login(SMTP_USER, SMTP_PASS)
            server.send_message(msg)

        print("メール送信成功")

    except Exception as e:
        print("メール送信失敗:", e)

def mail_new(data, time, name, phone):
    send_admin_mail(
        "【新規予約】",
        f"""
新規予約が入りました

--------------------
予約日付：{data}
予約時間：{format_time_range(time)}
氏名：{name}
電話：{phone}
--------------------
"""
    )


def mail_edit(data, time, name, phone):
    send_admin_mail(
        "【予約変更】",
        f"""
予約が変更されました

--------------------
予約日付：{data}
予約時間：{format_time_range(time)}
氏名：{name}
電話：{phone}
--------------------
"""
    )


def mail_delete(data, time, name, phone):
    send_admin_mail(
        "【予約削除】",
        f"""
予約が削除されました

--------------------
予約日付：{data}
予約時間：{format_time_range(time)}
氏名：{name}
電話：{phone}
--------------------
"""
    )

def send_mail(row):
    SMTP_USER = os.getenv("SMTP_USER")
    SMTP_PASS = os.getenv("SMTP_PASS")

    name, data, time, email = row

    print("RAW TIME =", time)

    try:
        t = time[:5]

        # 「09」みたいな場合の補正
        if ":" not in t:
            t = f"{t}:00"

        start = datetime.strptime(t, "%H:%M")
        end = start + timedelta(minutes=30)

        # ★ここ重要（必ず2桁表示）
        time_range = f"{start.strftime('%H:%M')}～{end.strftime('%H:%M')}"

    except:
        time_range = time or ""

    body = f"""{name} 様

ガス点検の予約が確定しました。

■日時
{data} {time_range}

上記の日時にお伺いしますので、ご在宅をお願いいたします。なお、都合により時間が前後する場合もございますが、ご容赦ください。
また、このメール受信以降に予約の変更を希望される際は、お手数ですがお電話にてご相談ください。
"""

    msg = MIMEText(body, "plain", "utf-8")

    # 件名（UTF-8）
    msg["Subject"] = str(Header("予約確定のお知らせ", "utf-8"))

    # ★ここも重要（FromもUTF-8にする）
    SMTP_USER = os.getenv("SMTP_USER")

    msg["From"] = formataddr(
        (str(Header("ガス点検", "utf-8")), SMTP_USER)
    )

    msg["To"] = email

    smtp = smtplib.SMTP_SSL("smtp.gmail.com", 465)
    smtp.login(SMTP_USER, SMTP_PASS)
    smtp.send_message(msg)
    smtp.quit()
# ---------------- 確定メール送信ここまで ----------------

from datetime import datetime, timedelta, timezone
JST = timezone(timedelta(hours=9))

app.secret_key = os.getenv("SECRET_KEY")
app.permanent_session_lifetime = timedelta(days=7)

# ---------------- トップ ----------------
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        code = request.form.get('consumer_code')

        if not code:
            return "消費者コードを入力してください"

        session['code'] = code
        action = request.form.get('action')

        if action == "新規":
            return redirect('/new')
        elif action == "変更":
            return redirect('/edit')
        elif action == "削除":
            return redirect('/delete')
        elif action == "確認":   # ★ここに追加
            return redirect('/view')

    return render_template('index.html')

# ---------------- ログイン ----------------

ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':

        pw = request.form.get('password')
        print("PASSWORD INPUT =", repr(pw))  # ←ここ

        ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")

        if pw == ADMIN_PASSWORD:
            session['login'] = True
            return redirect('/admin_menu')

        return render_template('login.html', error="パスワードが違います")

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

# ---------------- 管理者メイン画面 ----------------
@app.route('/admin_menu')
def admin_menu():
    if not session.get('login'):
        return redirect('/login')

    # -----------------------------
    # ① 予約ブロック一覧（今のまま維持）
    # -----------------------------
    res = supabase.table("blocked_times") \
        .select("id,data,start_time,end_time") \
        .order("data", desc=False) \
        .execute()

    blocks = res.data or []

    # -----------------------------
    # ② 予約可能期間（settingsから取得）
    # -----------------------------
    setting_res = supabase.table("settings") \
        .select("start_data,end_data") \
        .eq("id", 1) \
        .limit(1) \
        .execute()

    setting = setting_res.data[0] if setting_res.data else None

    start_data = setting.get("start_data") if setting else None
    end_data = setting.get("end_data") if setting else None

    # -----------------------------
    # ③ 画面へ
    # -----------------------------
    return render_template(
        'admin_menu.html',
        blocks=blocks,
        start_data=start_data,
        end_data=end_data
    )

# ---------------- 予約可能期間設定 ----------------
@app.route('/admin_setting')
def admin_setting():
    if not session.get('login'):
        return redirect('/login')

    res = supabase.table("settings") \
        .select("*") \
        .eq("id", 1) \
        .execute()

    setting = res.data[0] if res.data else None

    start = setting.get("start_data", "") if setting else ""
    end = setting.get("end_data", "") if setting else ""

    return render_template(
        "admin_setting.html",
        start=start,
        end=end
    )


@app.route('/save_setting', methods=['POST'])
def save_setting():

    start_data = request.form.get('start_data')
    end_data = request.form.get('end_data')

    if not start_data or not end_data:
        return redirect('/admin_setting')

    # ---------------- ログインチェック
    if not session.get('login'):
        return redirect('/login')

    # ---------------- フォーム取得
    start_data = request.form.get('start_data')
    end_data = request.form.get('end_data')

    # ---------------- Supabase保存（UPSERT）
    supabase.table("settings").upsert({
        "id": 1,
        "start_data": start_data,
        "end_data": end_data
    }).execute()

    # ---------------- 完了
    return redirect('/admin_menu')

@app.route('/clear_setting', methods=['POST'])
def clear_setting():

    if not session.get('login'):
        return redirect('/login')

    supabase.table("settings").upsert({
        "id": 1,
        "start_data": None,
        "end_data": None
    }).execute()

    return redirect('/admin_menu')

# ---------------- admin（★ここ改修） ----------------
@app.route('/admin')
def admin():
    if not session.get('login'):
        return redirect('/login')

    code = request.args.get('code', '')
    name = request.args.get('name', '')
    confirmed = request.args.get('confirmed', '')

    data_from = request.args.get('data_from', '')
    data_to = request.args.get('data_to', '')

    created_from = request.args.get('created_from', '')
    created_to = request.args.get('created_to', '')

    query = supabase.table("reservations").select("*").eq("is_deleted", False)

    # フィルター（Supabase版）
    if confirmed != "":
        query = query.eq("is_confirmed", confirmed == "1")

    if code:
        query = query.ilike("consumer_code", f"%{code}%")

    if name:
        query = query.ilike("name", f"%{name}%")

    if data_from:
        query = query.gte("data", data_from)

    if data_to:
        query = query.lte("data", data_to)

    if created_from:
        query = query.gte("created_at", created_from)

    if created_to:
        query = query.lte("created_at", created_to)

    res = query = query.order("consumer_code", desc=False).order("created_at", desc=True).execute()

    reservations = res.data or []

    for r in reservations:
        r["status"] = r.get("status") or "新規"

    return render_template("admin.html", reservations=reservations)

@app.route('/toggle_confirm', methods=['POST'])
def toggle_confirm():

    reservation_id = request.form.get('id')
    confirmed = request.form.get('confirmed') == 'on'

    if not reservation_id or not reservation_id.isdigit():
        return redirect('/admin')

    reservation_id = int(reservation_id)

    # ---------------- 更新
    supabase.table("reservations") \
        .update({"is_confirmed": confirmed}) \
        .eq("id", reservation_id) \
        .execute()

    # ---------------- 1件取得（安全版）
    res = supabase.table("reservations") \
        .select("name, data, time, email") \
        .eq("id", reservation_id) \
        .limit(1) \
        .execute()

    if not res.data:
        return redirect('/admin')

    r = res.data[0]

    row = (
        r.get("name"),
        r.get("data"),
        r.get("time"),
        r.get("email")
    )

    # ---------------- メール送信
    if confirmed and row and row[3]:
        send_mail(row)

    return redirect('/admin')

@app.route('/admin_delete', methods=['POST'])
def admin_delete():

    if not session.get('login'):
        return redirect('/login')

    reservation_id = request.form.get('id')

    if not reservation_id:
        return redirect('/admin')

    # -----------------------
    # 元データ取得
    # -----------------------
    r = supabase.table("reservations") \
        .select("*") \
        .eq("id", int(reservation_id)) \
        .limit(1) \
        .execute()

    if not r.data:
        return redirect('/admin')

    row = r.data[0]

    # -----------------------
    # ★ここが「移動」の本体
    # -----------------------
    supabase.table("reservations") \
        .update({
            "is_deleted": True
        }) \
        .eq("id", int(reservation_id)) \
        .execute()

    return redirect('/admin')

@app.route('/admin_edit/<int:id>')
def admin_edit(id):
    if not session.get('login'):
        return redirect('/login')

    # ----------------------------
    # Supabaseから取得（SQLite廃止）
    # ----------------------------
    res = supabase.table("reservations") \
        .select("id,data,time,name,phone,address,email") \
        .eq("id", id) \
        .execute()

    data = None

    if res.data:
        r = res.data[0]

        # Flaskテンプレ互換（タプル形式にする）
        data = (
            r.get("id"),
            r.get("data"),
            r.get("time"),
            r.get("name"),
            r.get("phone"),
            r.get("address"),
            r.get("email")
        )

    return render_template("admin_edit.html", data=data)

# ---------------- edit_save ---------------
@app.route('/edit_save', methods=['POST'])
def edit_save():

    data = request.form
    code = session.get('code')

    if not code:
        return redirect('/')

    # ---------------- 重複チェック（他人のみ）
    check = supabase.table("reservations") \
        .select("id") \
        .eq("data", data['data']) \
        .eq("time", data['time'][:5]) \
        .eq("is_deleted", False) \
        .neq("consumer_code", code) \
        .execute()

    if check.data:
        return "この時間は予約できません"

    # ---------------- 新しい履歴を追加
    supabase.table("reservations").insert({
        "data": data['data'],
        "time": data['time'][:5],
        "consumer_code": code,
        "name": data['name'],
        "phone": data['phone'],
        "address": data['address'],
        "email": data['email'],
        "status": "変更",
        "is_deleted": False,
        "is_confirmed": False,
        "created_at": datetime.now(JST).isoformat()
    }).execute()

    # ---------------- メール
    mail_edit(
        data['data'],
        data['time'][:5],
        data['name'],
        data['phone']
    )

    return render_template("edit_complete.html", data=data)

@app.route('/admin_deleted')
def admin_deleted():
    if not session.get('login'):
        return redirect('/login')

    # フィルタ取得
    data_from = request.args.get('data_from', '')
    data_to = request.args.get('data_to', '')
    name = request.args.get('name', '')
    code = request.args.get('code', '')

    # Supabaseクエリ
    query = supabase.table("reservations") \
        .select("*") \
        .eq("is_deleted", True)

    if data_from:
        query = query.gte("data", data_from)

    if data_to:
        query = query.lte("data", data_to)

    if name:
        query = query.ilike("name", f"%{name}%")

    if code:
        query = query.ilike("consumer_code", f"%{code}%")

    res = query.order("created_at", desc=True).execute()

    print("🔥 admin_deleted 生データ")
    print(res.data)

    rows = res.data or []

    # 時間変換
    new_rows = []
    for r in rows:
        try:
            t = r["time"][:5]
            h, m = map(int, t.split(":"))

            end_h = h
            end_m = m + 30
            if end_m >= 60:
                end_h += 1
                end_m -= 60

            time_range = f"{t}～{str(end_h).zfill(2)}:{str(end_m).zfill(2)}"
        except:
            time_range = r.get("time", "")

        new_rows.append({
            "id": r.get("id"),
            "data": r.get("data"),
            "time": time_range,
            "created_at": r.get("created_at"),
            "consumer_code": r.get("consumer_code"),
            "name": r.get("name"),
            "phone": r.get("phone"),
            "address": r.get("address"),
            "email": r.get("email"),
            "status":r.get("status"),
            "is_deleted": True
        })

    return render_template("admin_deleted.html", reservations=new_rows)

@app.route('/admin_restore', methods=['POST'])
def admin_restore():

    if not session.get('login'):
        return redirect('/login')

    reservation_id = request.form.get('id')

    # ★重要：actionは絶対触らない
    supabase.table("reservations") \
        .update({
            "is_deleted": False
        }) \
        .eq("id", int(reservation_id)) \
        .execute()

    return redirect('/admin_deleted')

@app.route('/admin_restore_multi', methods=['POST'])
def admin_restore_multi():
    if not session.get('login'):
        return redirect('/login')

    ids = request.form.getlist('ids')

    for i in ids:
        supabase.table("reservations") \
            .update({
                "is_deleted": False
            }) \
            .eq("id", int(i)) \
            .execute()

    return redirect('/admin_deleted')


@app.route('/admin_bulk_delete', methods=['POST'])
def admin_bulk_delete():

    if not session.get('login'):
        return redirect('/login')

    ids = request.form.getlist('ids')

    if not ids:
        return redirect('/admin_deleted')

    for i in ids:
        supabase.table("reservations") \
            .delete() \
            .eq("id", int(i)) \
            .execute()

    return redirect('/admin_deleted')

# ---------------- admin_block ----------------
@app.route('/admin_block')
def admin_block():
    if not session.get('login'):
        return redirect('/login')

    res = supabase.table("blocked_times") \
        .select("id,data,start_time,end_time") \
        .order("data", desc=False) \
        .execute()

    rows = res.data or []

    return render_template("admin_block.html", blocks=rows)

@app.route('/export_block_excel')
def export_block_excel():
    if not session.get('login'):
        return redirect('/login')

    res = supabase.table("blocked_times") \
        .select("id,data,start_time,end_time") \
        .order("data", desc=False) \
        .execute()

    rows = res.data or []

    wb = Workbook()
    ws = wb.active
    ws.title = "予約ブロック"

    ws.append(["日付", "開始", "終了"])

    for r in rows:
        ws.append([
        r.get("data"),
        r.get("start_time"),
        r.get("end_time")
    ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="reservation_blocks.xlsx",
        as_attachment=True
    )

@app.route('/add_block', methods=['POST'])
def add_block():

    # ---------------- ログインチェック
    if not session.get('login'):
        return redirect('/login')

    try:
        # ---------------- フォーム取得
        data = request.form.get('data')
        start_time = request.form.get('start_time')
        end_time = request.form.get('end_time')

        print("受信:", data, start_time, end_time)  # ★デバッグ

        # ---------------- 入力チェック
        if not data or not start_time or not end_time:
            print("入力不足")
            return redirect('/admin_block')

        start_time = start_time[:5]
        end_time = end_time[:5]

        # ---------------- 時間チェック
        start_dt = datetime.strptime(start_time, "%H:%M")
        end_dt = datetime.strptime(end_time, "%H:%M")

        if end_dt <= start_dt:
            print("時間NG")
            return redirect('/admin_block')

        # ---------------- 重複チェック
        existing = supabase.table("blocked_times") \
            .select("id") \
            .eq("data", data) \
            .eq("start_time", start_time) \
            .eq("end_time", end_time) \
            .execute()

        if existing.data:
            print("完全一致ブロックあり")
            return redirect('/admin_block')

        # ---------------- 予約との重複チェック
        reservations = supabase.table("reservations") \
            .select("time") \
            .eq("data", data) \
            .eq("is_deleted", False) \
            .execute()

        reserved_times = set()
        for r in (reservations.data or []):
            if r.get("time"):
                reserved_times.add(r["time"][:5])

        current = start_dt
        while current < end_dt:
            t = current.strftime("%H:%M")

            if t in reserved_times:
                print("予約と衝突:", t)
                return redirect('/admin_block')

            current += timedelta(minutes=30)

        # ---------------- ★ここ重要（結果確認）
        result = supabase.table("blocked_times").insert({
            "data": data,
            "start_time": start_time,
            "end_time": end_time
        }).execute()

        print("INSERT結果:", result)

        # ★失敗チェック
        if not result.data:
            print("INSERT失敗")
            return redirect('/admin_block')

        print("追加成功")

        return redirect('/admin_block')

    except Exception as e:
        print("エラー発生:", e)
        return redirect('/admin_block')


@app.route('/delete_block/<int:block_id>')
def delete_block(block_id):

    if not session.get('login'):
        return redirect('/login')

    supabase.table("blocked_times") \
        .delete() \
        .eq("id", block_id) \
        .execute()

    return redirect('/admin_block')

# ---------------- new ----------------
@app.route('/new', methods=['GET', 'POST'])
def new():
    if not session.get('code'):
        return redirect('/')

    # ★予約可能期間取得
    res = supabase.table("settings") \
        .select("*") \
        .eq("id", 1) \
        .limit(1) \
        .execute()

    setting = res.data[0] if res.data else None

    start_data = setting.get("start_data") if setting else ""
    end_data = setting.get("end_data") if setting else ""

    # ★デフォルト
    data = {
        "data": "",
        "time": "",
        "name": "",
        "phone": "",
        "address": "",
        "email": ""
    }

    # ★POST時は復元
    if request.method == 'POST':
        data = request.form

    return render_template(
        "new.html",
        data=data,
        start_data=start_data,
        end_data=end_data
    )

# ---------------- get_times ----------------
@app.route('/get_times')
def get_times():

    data = request.args.get('data')
    if not data:
        return jsonify([])

    # =========================
    # ① settings取得（Supabase化）
    # =========================
    setting_res = supabase.table("settings") \
        .select("start_data,end_data") \
        .eq("id", 1) \
        .limit(1) \
        .execute()

    setting = setting_res.data[0] if setting_res.data else None

    if setting:
        start_data = setting.get("start_data")
        end_data = setting.get("end_data")

        if start_data and data < start_data:
            return jsonify([])

        if end_data and data > end_data:
            return jsonify([])

    # =========================
    # ② 予約取得（Supabase）
    # =========================
    res = supabase.table("reservations") \
        .select("time") \
        .eq("data", data) \
        .eq("is_deleted", False) \
        .execute()

    reserved = set()
    for r in (res.data or []):
        if r.get("time"):
            reserved.add(r["time"][:5])

    # =========================
    # ③ ブロック取得（Supabase化）
    # =========================
    block_res = supabase.table("blocked_times") \
        .select("start_time,end_time") \
        .eq("data", data) \
        .execute()

    blocks = []
    for r in (block_res.data or []):
        blocks.append((r["start_time"], r["end_time"]))

    # =========================
    # ④ 時間生成
    # =========================
    limit_time = datetime.now(JST) + timedelta(hours=24)

    slots = []

    start = datetime.strptime("09:30", "%H:%M")
    end = datetime.strptime("16:30", "%H:%M")

    while start <= end:

        t = start.strftime("%H:%M")

        current_dt = datetime.strptime(
            data + " " + t,
            "%Y-%m-%d %H:%M"
        ).replace(tzinfo=JST)

        # ①予約済みチェック
        if t in reserved:
            start += timedelta(minutes=30)
            continue

        # ②ブロックチェック
        blocked = False

        for b_start, b_end in blocks:

            bs = datetime.strptime(data + " " + b_start, "%Y-%m-%d %H:%M").replace(tzinfo=JST)
            be = datetime.strptime(data + " " + b_end, "%Y-%m-%d %H:%M").replace(tzinfo=JST)

            # 日跨ぎ対応
            if be <= bs:
                be += timedelta(days=1)

            if bs <= current_dt < be:
                blocked = True
                break

        if blocked:
            start += timedelta(minutes=30)
            continue

        # ③24時間ルール
        #if current_dt < limit_time:
        #    start += timedelta(minutes=30)
        #    continue

        # OK
        slots.append(t)
        start += timedelta(minutes=30)

    return jsonify(slots)


@app.route('/confirm', methods=['POST'])
def confirm():

    data = {
        "data": request.form.get("data"),
        "time": request.form.get("time"),
        "name": request.form.get("name"),
        "phone": request.form.get("phone"),
        "address": request.form.get("address"),
        "email": request.form.get("email")
    }

    return render_template("confirm.html", data=data)

# ---------------- create ----------------
@app.route('/create_confirm', methods=['POST'])
def create_confirm():

    data = request.form  # ←先にこれ
    print("FORM DATE =", dict(data))  # ←ここでOK

    code = session.get('code')

    if not code:
        return "ログイン情報なし"

    # ---------------- 予約時間チェック ----------------
    target_dt = datetime.strptime(
        data['data'] + " " + data['time'][:5],
        "%Y-%m-%d %H:%M"
    ).replace(tzinfo=JST)

    limit_time = datetime.now(JST) + timedelta(hours=24)

    if target_dt < limit_time:
        return "24時間後以降の予約しかできません"

# ---------------- 重複チェック ----------------
    check = supabase.table("reservations") \
        .select("id") \
        .eq("data", data.get("data")) \
        .eq("time", data.get("time")[:5]) \
        .eq("is_deleted", False) \
        .execute()

    if check.data:
        return "この時間は予約できません。"

    # ---------------- Supabase保存（競合対策） ----------------
    try:
        supabase.table("reservations").insert({
            "data": data.get("data", ""),
            "time": data.get("time", "")[:5],
            "consumer_code": code,
            "name": data.get("name", ""),
            "phone": data.get("phone", ""),
            "address": data.get("address", ""),
            "email": data.get("email", ""),
            "status": "新規",
            "is_deleted": False,
            "is_confirmed": False,
            "is_deleted": False,
            "is_confirmed": False,
            "created_at": datetime.now(JST).isoformat()
        }).execute()

    except Exception as e:
        print("INSERT ERROR =", e)
        return "この時間はすでに予約されています"

    # ---------------- メール送信 ----------------
    mail_new(
        data.get('data', ''),
        data.get('time', '')[:5],
        data.get('name', ''),
        data.get('phone', '')
    )

    return render_template("complete.html", data=data)

# ---------------- check_day ----------------
@app.route('/check_day')
def check_day():

    data = request.args.get('data')
    if not data:
        return jsonify({"ok": False})

    # -------------------------------
    # blocked_times（Supabase取得）
    # -------------------------------
    res = supabase.table("blocked_times") \
        .select("start_time,end_time") \
        .eq("data", data) \
        .execute()

    blocks = res.data or []

    # -------------------------------
    # 予約枠（固定）
    # -------------------------------
    slots = []
    start = datetime.strptime("09:30", "%H:%M")
    end = datetime.strptime("16:30", "%H:%M")

    while start <= end:
        slots.append(start.strftime("%H:%M"))
        start += timedelta(minutes=30)

    # -------------------------------
    # ブロック削除ロジック
    # -------------------------------
    for b in blocks:

        b_start = b.get("start_time")
        b_end = b.get("end_time")

        if not b_start or not b_end:
            continue

        bs = datetime.strptime(b_start, "%H:%M")
        be = datetime.strptime(b_end, "%H:%M")

        slots = [
            t for t in slots
            if not (bs <= datetime.strptime(t, "%H:%M") < be)
        ]

    # -------------------------------
    # 結果
    # -------------------------------
    return jsonify({
        "ok": len(slots) > 0,
        "message": "この日は予約できません" if len(slots) == 0 else ""
    })

# ---------------- edit ----------------
@app.route('/edit')
def edit():
    code = session.get('code')

    if not code:
        return redirect('/')

    # ----------------------------
    # Supabaseから予約取得
    # ----------------------------
    res = supabase.table("reservations") \
        .select("*") \
        .eq("consumer_code", code) \
        .eq("is_deleted", False) \
        .order("id", desc=True) \
        .limit(1) \
        .execute()

    data = res.data[0] if res.data else None

    if not data:
        return "予約データがありません"

    # ----------------------------
    # 期間取得
    # ----------------------------
    res = supabase.table("settings") \
        .select("*") \
        .eq("id", 1) \
        .limit(1) \
        .execute()

    setting = res.data[0] if res.data else None

    start_data = setting.get("start_data") if setting else ""
    end_data = setting.get("end_data") if setting else ""

    return render_template(
        "edit.html",
        data=(
            data["id"],
            data["data"],
            data["time"],
            data["name"],
            data["phone"],
            data["address"],
            data.get("email", "")
        ),
        start_data=start_data,
        end_data=end_data
    )

@app.route('/edit_confirm', methods=['POST'])
def edit_confirm():

    print(request.form)

    data = {
        "data": request.form.get("data"),
        "time": request.form.get("time"),
        "name": request.form.get("name"),
        "phone": request.form.get("phone"),
        "address": request.form.get("address"),
        "email": request.form.get("email")
    }

    return render_template("edit_confirm.html", data=data)

# ---------------- delete ----------------
# -------------------------------
# 削除確認画面（表示だけ）
# -------------------------------
@app.route('/delete')
def delete():
    code = session.get('code')

    if not code:
        return redirect('/')

    res = supabase.table("reservations") \
        .select("*") \
        .eq("consumer_code", code) \
        .eq("is_deleted", False) \
        .order("id", desc=True) \
        .limit(1) \
        .execute()

    row = res.data[0] if res.data else None

    if not row:
        return render_template("delete.html", data=None)

    start = datetime.strptime(row["time"][:5], "%H:%M")
    end = start + timedelta(minutes=30)

    data = (
        row["id"],
        row["data"],
        f"{start:%H:%M}～{end:%H:%M}",
        row["name"],
        row["phone"],
        row["address"],
        row.get("email", "")
    )

    return render_template("delete.html", data=data)


# -------------------------------
# 削除実行（ここがPOST）
# -------------------------------
@app.route('/delete_confirm', methods=['POST'])
def delete_confirm():

    code = session.get('code')

    if not code:
        return redirect('/')

    # 最新予約取得
    res = supabase.table("reservations") \
        .select("*") \
        .eq("consumer_code", code) \
        .eq("is_deleted", False) \
        .order("created_at", desc=True) \
        .limit(1) \
        .execute()

    if not res.data:
        return redirect('/')

    r = res.data[0]

    # 削除履歴を追加
    supabase.table("reservations").insert({
        "data": r["data"],
        "time": r["time"],
        "consumer_code": r["consumer_code"],
        "name": r["name"],
        "phone": r["phone"],
        "address": r["address"],
        "email": r.get("email"),
        "status": "削除",
        "is_deleted": False,
        "is_confirmed": False,
        "created_at": datetime.now(JST).isoformat()
    }).execute()

    # ★これ追加
    mail_delete(
        r["data"],
        r["time"][:5],
        r["name"],
        r["phone"]
    )

    return render_template("delete_done.html")

# ---------------- 予約内容確認 ----------------
@app.route('/view')
def view():
    code = session.get('code')

    if not code:
        return redirect('/')

    res = supabase.table("reservations") \
        .select("*") \
        .eq("consumer_code", code) \
        .eq("is_deleted", False) \
        .order("id", desc=True) \
        .limit(1) \
        .execute()

    row = res.data[0] if res.data else None

    # ② rowがない場合
    if not row:
        return render_template("view.html", data=None)

    # ③ time整形
    try:
        start = datetime.strptime(row["time"][:5], "%H:%M")
        end = start + timedelta(minutes=30)
        time_range = f"{start.strftime('%H:%M')}～{end.strftime('%H:%M')}"
    except:
        time_range = row[2] or ""

    # ④ ここでdata作る（←ここでrow使うのが正解）
    data = (
        row["id"],
        row["data"],
        time_range,
        row["name"],
        row["phone"],
        row["address"],
        row.get("email"),
        row.get("is_confirmed")
    )

    return render_template("view.html", data=data)

# ---------------- 予約excel出力 ----------------
@app.route('/export_excel')
def export_excel():
    if not session.get('login'):
        return redirect('/login')

    # Supabaseから取得
    res = supabase.table("reservations") \
        .select("*") \
        .eq("is_deleted", False) \
        .order("created_at", desc=True) \
        .execute()

    rows = res.data or []

    # Excel作成
    wb = Workbook()
    ws = wb.active
    ws.title = "予約一覧"

    # ヘッダー
    ws.append([
        "予約日", "時間", "申込日時",
        "コード", "氏名", "住所",
        "電話", "メール", "状態"
    ])

    for r in rows:
        # 時間整形
        time = (r.get("time") or "")[:5]

        # 日時整形
        created = (r.get("created_at") or "")
        if created:
            created = created[:19].replace("T", " ")

        ws.append([
            r.get("data"),
            time,
            created,
            r.get("consumer_code"),
            r.get("name"),
            r.get("address"),
            r.get("phone"),
            r.get("email"),
            r.get("status") or "new"
        ])

    # 出力
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="reservations.xlsx",
        as_attachment=True
    )

# ---------------- 予約削除excel出力 ----------------
@app.route('/export_deleted_excel')
def export_deleted_excel():

    if not session.get('login'):
        return redirect('/login')

    res = supabase.table("reservations") \
        .select("*") \
        .eq("is_deleted", True) \
        .order("created_at", desc=True) \
        .execute()

    rows = res.data or []

    wb = Workbook()
    ws = wb.active
    ws.title = "削除一覧"

    ws.append([
        "予約日",
        "予約時間",
        "申込日時",
        "消費者コード",
        "氏名",
        "住所",
        "電話番号",
        "状態"
    ])

    for r in rows:
        time = (r.get("time") or "")[:5]

        created = (r.get("created_at") or "")
        if created:
            created = created[:19].replace("T", " ")

        ws.append([
            r.get("data"),
            time,
            created,
            r.get("consumer_code"),
            r.get("name"),
            r.get("address"),
            r.get("phone"),
            r.get("status") or "new"
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="deleted_reservations.xlsx",
        as_attachment=True
    )

# ---------------- 起動 ----------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000, debug=True)